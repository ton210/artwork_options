#!/usr/bin/env python3
"""
Create detailed state-by-state customer lists
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("Creating detailed state-by-state customer lists...")

# Load the analysis data
excel_file = 'MunchMakers_Sales_Analysis.xlsx'
major_buyers = pd.read_excel(excel_file, sheet_name='Major Buyers >$300')
state_analysis = pd.read_excel(excel_file, sheet_name='Analysis by State')

# Sort major buyers by state and then by revenue
major_buyers_sorted = major_buyers.sort_values(['state', 'total_revenue'], ascending=[True, False])

# Create new workbook with state lists
output_file = 'MunchMakers_Buyers_By_State.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

    # Summary sheet - all states with totals
    summary_data = []
    for _, state_row in state_analysis.iterrows():
        state = state_row['state'] if pd.notna(state_row['state']) and str(state_row['state']).strip() else 'Unknown'
        state_buyers = major_buyers[major_buyers['state'] == state_row['state']]

        summary_data.append({
            'State': state,
            'Total Revenue': state_row['revenue'],
            'Customer Count': int(state_row['customer_count']),
            'Total Orders': int(state_row['total_orders']),
            'Avg Order Value': state_row['avg_order_value']
        })

    summary_df = pd.DataFrame(summary_data).sort_values('Total Revenue', ascending=False)
    summary_df.to_excel(writer, sheet_name='Summary by State', index=False)

    # Individual sheets for each state (top 30 states)
    for idx, state_row in state_analysis.head(30).iterrows():
        state = state_row['state'] if pd.notna(state_row['state']) and str(state_row['state']).strip() else 'Unknown'
        state_buyers = major_buyers[major_buyers['state'] == state_row['state']].copy()
        state_buyers = state_buyers.sort_values('total_revenue', ascending=False)

        # Clean up sheet name
        sheet_name = str(state)[:31]  # Excel sheet name limit
        if not sheet_name:
            sheet_name = 'Unknown'

        # Select relevant columns
        state_buyers_export = state_buyers[['email', 'name', 'company', 'total_revenue', 'total_orders']].copy()
        state_buyers_export.columns = ['Email', 'Name', 'Company', 'Total Revenue', 'Total Orders']

        state_buyers_export.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"✓ Created: {output_file}")

# Also create a US-only focused list
us_buyers = major_buyers[major_buyers['country'] == 'US'].copy()
us_buyers_sorted = us_buyers.sort_values('total_revenue', ascending=False)

with pd.ExcelWriter('MunchMakers_US_Buyers.xlsx', engine='openpyxl') as writer:
    # All US buyers
    us_export = us_buyers_sorted[['email', 'name', 'company', 'state', 'total_revenue', 'total_orders']].copy()
    us_export.columns = ['Email', 'Name', 'Company', 'State', 'Total Revenue', 'Total Orders']
    us_export.to_excel(writer, sheet_name='All US Buyers', index=False)

    # Top 100 US buyers
    us_top100 = us_export.head(100)
    us_top100.to_excel(writer, sheet_name='Top 100 US Buyers', index=False)

print(f"✓ Created: MunchMakers_US_Buyers.xlsx")

# International buyers (non-US)
intl_buyers = major_buyers[major_buyers['country'] != 'US'].copy()
intl_buyers = intl_buyers[intl_buyers['country'].notna()]  # Remove empty country
intl_buyers_sorted = intl_buyers.sort_values('total_revenue', ascending=False)

with pd.ExcelWriter('MunchMakers_International_Buyers.xlsx', engine='openpyxl') as writer:
    # All international buyers
    intl_export = intl_buyers_sorted[['email', 'name', 'company', 'country', 'total_revenue', 'total_orders']].copy()
    intl_export.columns = ['Email', 'Name', 'Company', 'Country', 'Total Revenue', 'Total Orders']
    intl_export.to_excel(writer, sheet_name='All International', index=False)

    # By country
    for country in intl_buyers['country'].unique()[:20]:  # Top 20 countries
        country_buyers = intl_buyers[intl_buyers['country'] == country].copy()
        country_buyers = country_buyers.sort_values('total_revenue', ascending=False)

        sheet_name = str(country)[:31]
        country_export = country_buyers[['email', 'name', 'company', 'total_revenue', 'total_orders']].copy()
        country_export.columns = ['Email', 'Name', 'Company', 'Total Revenue', 'Total Orders']
        country_export.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"✓ Created: MunchMakers_International_Buyers.xlsx")

print("\n" + "=" * 80)
print("STATE-BY-STATE ANALYSIS COMPLETE")
print("=" * 80)
print(f"Files created:")
print(f"  - MunchMakers_Buyers_By_State.xlsx (30 states)")
print(f"  - MunchMakers_US_Buyers.xlsx (US-focused analysis)")
print(f"  - MunchMakers_International_Buyers.xlsx (International customers)")
print(f"\nTotal major buyers analyzed: {len(major_buyers)}")
print(f"  - US buyers: {len(us_buyers)}")
print(f"  - International buyers: {len(intl_buyers)}")
